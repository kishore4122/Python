import openpyxl

#Loading the Spreadsheetfile and save it to a variable
inv_file = openpyxl.load_workbook("c:\\Users\\chand\\Desktop\\DevOps\\Materials\\Python\\TechWorld-Nana\\Projects\\Spreadsheetfiles\\inventory.xlsx")

#Read the specific sheet from spreadsheet file and save it to variable
product_list = inv_file["Sheet1"]


products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# print(product_list.max_row) print the number of rows in a sheet

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value 
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_no = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # Calculation number of products per supplier
    # Task:1 - Calculate how many products from supplier
    # List each company with product count
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name) # her we get the supplier names and save them to a variable
        products_per_supplier[supplier_name] = current_num_products + 1
        #products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1 
    else:
        products_per_supplier[supplier_name] = 1   


    # Calculation total value of inventory per supplier
    # List each company with respective total inventory value
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_no)] = int(inventory)

    # Add value for Total Inventory Price
    # Write to Spreadsheet calculate and write inventory value for each product into spreadsheet
    inventory_price.value = inventory * price

    

print(products_per_supplier) 
print(total_value_per_supplier)
print(products_under_10_inv)
inv_file.save("c:\\Users\\chand\\Desktop\\DevOps\\Materials\\Python\\TechWorld-Nana\\Projects\\Spreadsheetfiles\\inventory_with_total_value.xlsx")
